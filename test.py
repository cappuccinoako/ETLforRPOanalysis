import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook

#TODO: create some func to import xlsx file and some gui idk
#import wb
wb = load_workbook(filename='Data\Machanical (Suspension).xlsx')


def readCritalParams(wb):
    """read Mechanical Critical Parameters from Mechanical excel file"""
    #HARDCODED
    cells = ['A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'F2', 'F3', 'F4', 'F5', 'F6', 'F7']
    values = [wb.active[cell].value for cell in cells]
    return(values)

def readInspectionData(wb):
    """read Inspection Data from Mechanical excel file"""
    #HARDCODED
    # ws = wb.active
    all_data = []
    columns_name = []
    ws = wb.active
    cell_range = ws['A9':'A59']
    for row in cell_range:
        for cell in row:
            columns_name.append(cell.value)
    for sheet in wb.worksheets:
        for column_range in range (4, 10):
            data = []
            for i in range(9, 60):
                if sheet.cell(column=column_range, row=9).value is None:
                    break
                data.append(sheet.cell(column=column_range, row=i).value)
            if data != []:
                all_data.append(data)
    df = pd.DataFrame(all_data, columns = columns_name) 
    return(df)