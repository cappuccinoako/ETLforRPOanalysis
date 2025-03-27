import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np

#TODO: create some func to import xlsx file and some gui idk
#import wb
wb = load_workbook(filename='/opt/airflow/data/Machanical (Suspension).xlsx')


def readCritalParams(wb):
    """read Mechanical Critical Parameters from Mechanical excel file"""
    #HARDCODED
    cells = ['A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'F2', 'F3', 'F4', 'F5', 'F6', 'F7']
    values = [wb.active[cell].value for cell in cells]
    return(values)

def readInspectionData(wb):
    """read Inspection Data from Mechanical excel file"""
    #HARDCODED
    # ws = wb.active
    all_data = []
    columns_name = []
    ws = wb.active
    cell_range = ws['A9':'A59']
    for row in cell_range:
        for cell in row:
            columns_name.append(cell.value)
    for sheet in wb.worksheets:
        for column_range in range (4, 10):
            data = []
            for i in range(9, 60):
                if sheet.cell(column=column_range, row=9).value is None:
                    break
                data.append(sheet.cell(column=column_range, row=i).value)
            if data != []:
                all_data.append(data)
    df = pd.DataFrame(all_data, columns = columns_name) 
    return(df)

def dataQuality(loaded_df):
    """check the data quality"""
    #check whether loaded_df is empty
    if loaded_df.empty:
        print('No Data Extracted')
        return False
    #replace None with NaN
    loaded_df = loaded_df.fillna(value=np.nan)
    return loaded_df

def rpo(exelFile=wb):
    #Importing the songs_df from the Extract.py
    load_df=readInspectionData(exelFile)
    dataQuality(load_df)
    return (load_df)

rpo()
